import csv
import io
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from database import get_db
from models.entities import Product, ProductCostHistory, User, WbToken
from routes.deps import current_user

router = APIRouter(prefix="/products", tags=["products"])


class CostPriceIn(BaseModel):
    cost_price: float | None
    tax_rate: float | None = None
    vendor_code: str = ""
    name: str = ""


def _active_token(db: Session, user_id: int) -> WbToken | None:
    return db.scalar(
        select(WbToken)
        .where(WbToken.user_id == user_id, WbToken.is_active.is_(True), WbToken.token_status != "invalid")
        .order_by(WbToken.id.desc())
    )


def _find_product_by_user_nm(db: Session, user_id: int, nm_id: int) -> Product | None:
    """Find product by user_id + nm_id across all tokens."""
    return db.scalar(
        select(Product)
        .where(Product.user_id == user_id, Product.nm_id == nm_id)
        .order_by(Product.id.desc())
    )


def _find_product_by_user_vendor(db: Session, user_id: int, vendor_code: str) -> Product | None:
    if not vendor_code:
        return None
    return db.scalar(
        select(Product)
        .where(Product.user_id == user_id, Product.vendor_code == vendor_code)
        .order_by(Product.id.desc())
    )


def _record_history(db: Session, user_id: int, product: Product, new_cost: float | None) -> None:
    if product.cost_price == new_cost:
        return
    db.add(ProductCostHistory(
        user_id=user_id,
        nm_id=product.nm_id,
        vendor_code=product.vendor_code or "",
        old_cost_price=product.cost_price,
        new_cost_price=new_cost,
    ))


@router.get("")
def list_products(user: User = Depends(current_user), db: Session = Depends(get_db)):
    products = db.scalars(
        select(Product)
        .where(Product.user_id == user.id)
        .order_by(Product.nm_id.asc())
    ).all()
    seen_nm: set[int] = set()
    result = []
    for p in products:
        if p.nm_id in seen_nm:
            continue
        seen_nm.add(p.nm_id)
        result.append({
            "id": p.id,
            "nm_id": p.nm_id,
            "vendor_code": p.vendor_code or "",
            "name": p.name or "",
            "cost_price": p.cost_price,
            "tax_rate": p.tax_rate,
            "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        })
    return result


@router.put("/{nm_id}/cost-price")
@router.patch("/{nm_id}/cost-price")
def update_cost_price(nm_id: int, payload: CostPriceIn, user: User = Depends(current_user), db: Session = Depends(get_db)):
    if payload.cost_price is not None and payload.cost_price < 0:
        raise HTTPException(status_code=400, detail="Себестоимость не может быть отрицательной")

    product = _find_product_by_user_nm(db, user.id, nm_id)
    if not product:
        wb_token = _active_token(db, user.id)
        product = Product(
            user_id=user.id,
            wb_token_id=wb_token.id if wb_token else None,
            nm_id=nm_id,
            vendor_code=payload.vendor_code or "",
            name=payload.name or "",
        )
        db.add(product)
        db.flush()

    _record_history(db, user.id, product, payload.cost_price)
    product.cost_price = payload.cost_price
    if payload.tax_rate is not None:
        product.tax_rate = payload.tax_rate
    if payload.vendor_code and not product.vendor_code:
        product.vendor_code = payload.vendor_code
    if payload.name and not product.name:
        product.name = payload.name
    db.commit()
    db.refresh(product)
    return {
        "ok": True,
        "id": product.id,
        "nm_id": product.nm_id,
        "vendor_code": product.vendor_code,
        "name": product.name,
        "cost_price": product.cost_price,
        "tax_rate": product.tax_rate,
        "updated_at": product.updated_at.isoformat() if product.updated_at else None,
    }


@router.post("/cost-prices/import")
async def import_cost_prices(
    file: UploadFile = File(...),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    file_name = file.filename or ""
    suffix = Path(file_name).suffix.lower()
    if suffix not in {".xlsx", ".xls", ".csv"}:
        raise HTTPException(status_code=400, detail="Поддерживаются файлы .xlsx, .xls, .csv")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    try:
        rows = _parse_import_file(suffix, content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}") from exc

    updated = 0
    not_found = 0
    errors: list[dict] = []
    wb_token = _active_token(db, user.id)

    for row_num, row in enumerate(rows, start=2):
        nm_id_raw = row.get("nm_id") or row.get("nmid") or row.get("nmID")
        vendor_code_raw = row.get("vendor_code") or row.get("артикул") or row.get("article") or ""
        cost_raw = row.get("cost_price") or row.get("себестоимость") or row.get("cost")

        nm_id = _parse_int(nm_id_raw)
        vendor_code = str(vendor_code_raw).strip() if vendor_code_raw else ""
        cost = _parse_float(cost_raw)

        if cost is None:
            errors.append({"row": row_num, "reason": "Неверное значение себестоимости"})
            continue
        if cost < 0:
            errors.append({"row": row_num, "reason": "Себестоимость не может быть отрицательной"})
            continue

        product = None
        if nm_id:
            product = _find_product_by_user_nm(db, user.id, nm_id)
        if not product and vendor_code:
            product = _find_product_by_user_vendor(db, user.id, vendor_code)

        if not product:
            if nm_id:
                product = Product(user_id=user.id, wb_token_id=wb_token.id if wb_token else None, nm_id=nm_id, vendor_code=vendor_code)
                db.add(product)
                db.flush()
            else:
                not_found += 1
                errors.append({"row": row_num, "reason": "Товар не найден (нет nmID или артикула)"})
                continue

        _record_history(db, user.id, product, cost)
        product.cost_price = cost
        updated += 1

    db.commit()
    return {
        "ok": True,
        "updated": updated,
        "not_found": not_found,
        "errors": errors,
        "message": f"Обновлено товаров: {updated}. Не найдено: {not_found}.",
    }


def _parse_import_file(suffix: str, content: bytes) -> list[dict]:
    if suffix == ".csv":
        return _parse_csv(content)
    return _parse_excel(suffix, content)


def _parse_csv(content: bytes) -> list[dict]:
    text = None
    for encoding in ("utf-8-sig", "cp1251", "utf-16"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось декодировать CSV")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [_normalize_row(row) for row in reader]


def _parse_excel(suffix: str, content: bytes) -> list[dict]:
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        return [_normalize_row(dict(zip(headers, row))) for row in rows[1:] if any(v is not None for v in row)]
    else:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        if ws.nrows < 2:
            return []
        headers = [str(ws.cell_value(0, c) or "").strip().lower() for c in range(ws.ncols)]
        result = []
        for r in range(1, ws.nrows):
            row_values = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if any(v for v in row_values):
                result.append(_normalize_row(dict(zip(headers, row_values))))
        return result


def _normalize_row(row: dict) -> dict:
    return {str(k or "").strip().lower(): v for k, v in row.items()}


def _parse_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _parse_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None
