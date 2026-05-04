import csv
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from models.entities import FinancialReport, FinancialReportItem, Product, WbToken

logger = logging.getLogger(__name__)


class FinancialReportError(ValueError):
    pass


class EmptyFinancialReport(FinancialReportError):
    pass


class MissingColumns(FinancialReportError):
    pass


class UnreadableFinancialReport(FinancialReportError):
    pass


EXPENSE_FIELDS = [
    "commission",
    "logistics",
    "storage",
    "returns",
    "acquiring",
    "spa",
    "advertising",
    "penalties",
    "deductions",
    "other_expenses",
]

REPORT_COLUMN_ALIASES = {
    "nm_id": ["Код номенклатуры", "nmID", "nmId"],
    "vendor_code": ["Артикул поставщика"],
    "product_name": ["Название"],
    "doc_type": ["Тип документа"],
    "operation": ["Обоснование для оплаты"],
    "quantity": ["Кол-во"],
    "sales_amount": ["Вайлдберриз реализовал Товар (Пр)", "Вайлдберриз реализовал Товар"],
    "retail_price": ["Цена розничная"],
    "to_pay": ["К перечислению Продавцу за реализованный Товар", "К перечислению продавцу"],
    "commission": ["Вознаграждение Вайлдберриз (ВВ), без НДС", "Вознаграждение Вайлдберриз"],
    "logistics": ["Услуги по доставке товара покупателю", "Логистика"],
    "storage": ["Хранение"],
    "penalties": ["Общая сумма штрафов", "Штрафы"],
    "deductions": ["Удержания"],
    "spa": ["Операции на приемке", "Операции на приёмке", "СПА"],
    "acquiring": [
        "Компенсация платёжных услуг/Комиссия за интеграцию платёжных сервисов",
        "Компенсация платежных услуг/Комиссия за интеграцию платежных сервисов",
        "Компенсация платёжных услуг",
        "Комиссия за интеграцию платёжных сервисов",
        "Эквайринг",
    ],
    "returns": ["Количество возврата", "Возвраты"],
    "delivery_count": ["Количество доставок"],
    "advertising": ["Реклама"],
    "tax": ["Налог"],
    "other_expenses": ["Прочие дополнительные расходы", "Прочие расходы"],
    "date": ["Дата продажи", "Дата операции", "Дата", "saleDt", "rrDate", "orderDt"],
}

REQUIRED_REPORT_FIELDS = [
    "nm_id",
    "vendor_code",
    "product_name",
    "operation",
    "quantity",
    "sales_amount",
    "retail_price",
    "to_pay",
    "commission",
    "logistics",
    "storage",
    "penalties",
    "deductions",
    "spa",
    "acquiring",
    "returns",
    "date",
]

VALIDATION_TOTALS = {
    "total_sales": "sales_amount",
    "total_for_pay": "to_pay",
    "total_commission": "commission",
    "total_logistics": "logistics",
    "total_storage": "storage",
    "total_penalties": "penalties",
    "total_deductions": "deductions",
    "total_acquiring": "acquiring",
    "total_spa": "spa",
    "total_returns": "returns",
    "total_other_expenses": "other_expenses",
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


NORMALIZED_ALIASES = {
    field: {_normalize_header(alias) for alias in aliases}
    for field, aliases in REPORT_COLUMN_ALIASES.items()
}


@dataclass
class ParsedReport:
    rows_count: int
    period_start: date | None
    period_end: date | None
    items: list[dict]
    source_rows: list[dict]
    columns_found: dict[str, str | None]
    validation: dict


def process_upload(db: Session, user_id: int, file_name: str, content: bytes) -> dict:
    parsed = parse_financial_report(file_name, content)
    if not parsed.items:
        raise EmptyFinancialReport("Файл пустой или не содержит данных.")

    _upsert_products_from_items(db, user_id, parsed.items)

    report = FinancialReport(
        user_id=user_id,
        file_name=file_name,
        period_start=parsed.period_start,
        period_end=parsed.period_end,
        rows_count=parsed.rows_count,
        products_count=len(parsed.items),
        manual_tax=None,
        validation_json=parsed.validation,
        source_rows_json=parsed.source_rows,
    )
    db.add(report)
    db.flush()

    item_rows = []
    for item in build_items(db, user_id, parsed.items, manual_tax=report.manual_tax):
        item_row = FinancialReportItem(report_id=report.id, user_id=user_id, **item)
        db.add(item_row)
        item_rows.append(item_row)
    db.commit()
    db.refresh(report)
    for item_row in item_rows:
        db.refresh(item_row)
    return report_payload(report, [_item_payload(item_row) for item_row in item_rows])


def validate_upload(file_name: str, content: bytes) -> dict:
    return parse_financial_report(file_name, content).validation


def parse_financial_report(file_name: str, content: bytes) -> ParsedReport:
    rows = _read_rows(file_name, content)
    if not rows:
        raise EmptyFinancialReport("Файл пустой или не содержит данных.")

    header_index, column_map = _detect_header(rows)
    header = rows[header_index]
    columns_found = _columns_found(header, column_map)
    missing = [field for field in REQUIRED_REPORT_FIELDS if field not in column_map]
    if missing:
        labels = ", ".join(REPORT_COLUMN_ALIASES[field][0] for field in missing)
        raise MissingColumns(f"Не найдены обязательные колонки финансового отчёта WB: {labels}.")

    data_rows = rows[header_index + 1 :]
    if not data_rows:
        raise EmptyFinancialReport("Файл пустой или не содержит данных.")

    period_start: date | None = None
    period_end: date | None = None
    source_rows = 0
    normalized_rows: list[dict] = []

    for row in data_rows:
        if _blank_row(row):
            continue
        source_rows += 1

        row_date = _date(_cell(row, column_map.get("date")))
        if row_date:
            if period_start is None or row_date < period_start:
                period_start = row_date
            if period_end is None or row_date > period_end:
                period_end = row_date

        nm_id = _int(_cell(row, column_map.get("nm_id")))
        vendor_code = _text(_cell(row, column_map.get("vendor_code")))
        product_name = _text(_cell(row, column_map.get("product_name")))
        if not nm_id and not vendor_code and not product_name:
            continue

        doc_type = _text(_cell(row, column_map.get("doc_type"))) if "doc_type" in column_map else ""
        operation = _text(_cell(row, column_map.get("operation")))
        quantity = _int(_cell(row, column_map.get("quantity"))) or 0
        delivery_count = _int(_cell(row, column_map.get("delivery_count"))) or 0
        retail_price = _money_value(_cell(row, column_map["retail_price"])) or 0.0
        sales_amount = _money_value(_cell(row, column_map["sales_amount"])) or 0.0
        to_pay = _money_value(_cell(row, column_map["to_pay"])) or 0.0
        commission = abs(_money_value(_cell(row, column_map["commission"])) or 0.0)
        logistics = abs(_money_value(_cell(row, column_map["logistics"])) or 0.0)
        storage = abs(_money_value(_cell(row, column_map["storage"])) or 0.0)
        penalties = abs(_money_value(_cell(row, column_map["penalties"])) or 0.0)
        deductions = abs(_money_value(_cell(row, column_map["deductions"])) or 0.0)
        acquiring = abs(_money_value(_cell(row, column_map.get("acquiring"))) or 0.0)
        spa = abs(_money_value(_cell(row, column_map.get("spa"))) or 0.0)
        returns = abs(_money_value(_cell(row, column_map.get("returns"))) or 0.0)
        advertising = _money_value(_cell(row, column_map.get("advertising"))) if "advertising" in column_map else None
        other_expenses = _money_value(_cell(row, column_map.get("other_expenses"))) if "other_expenses" in column_map else None
        is_sale = _is_sale_operation(operation, doc_type)

        normalized_rows.append(
            {
                "date": row_date.isoformat() if row_date else None,
                "nm_id": nm_id,
                "vendor_code": vendor_code,
                "product_name": product_name or (f"Товар {nm_id}" if nm_id else vendor_code),
                "doc_type": doc_type,
                "operation": operation,
                "is_sale": is_sale,
                "quantity": quantity,
                "delivery_count": delivery_count,
                "sold_qty": max(quantity, 0) if is_sale else 0,
                "retail_price": retail_price,
                "sales_amount": sales_amount,
                "before_spp": retail_price * max(quantity, 0) if is_sale else 0.0,
                "after_spp": sales_amount,
                "to_pay": to_pay,
                "commission": commission,
                "logistics": logistics,
                "storage": storage,
                "returns": returns,
                "acquiring": acquiring,
                "spa": spa,
                "advertising": advertising,
                "penalties": penalties,
                "deductions": deductions,
                "other_expenses": other_expenses,
            }
        )

    grouped = _group_report_rows(normalized_rows)
    validation = _build_validation(columns_found, source_rows, period_start, period_end, normalized_rows, grouped.values())
    logger.info(
        "Financial report parsed file=%s rows=%s products=%s period_start=%s period_end=%s status=%s diff=%s",
        file_name,
        source_rows,
        len(grouped),
        period_start,
        period_end,
        validation["status"],
        validation["diff"],
    )
    return ParsedReport(
        rows_count=source_rows,
        period_start=period_start,
        period_end=period_end,
        items=list(grouped.values()),
        source_rows=normalized_rows,
        columns_found=columns_found,
        validation=validation,
    )


def build_items(
    db: Session,
    user_id: int,
    parsed_items: list[dict],
    manual_tax: float | None = None,
    manual_advertising: float | None = None,
    manual_other_expenses: float | None = None,
) -> list[dict]:
    products = _product_lookup(db, user_id)
    result = []

    for source in parsed_items:
        nm_id = source["nm_id"]
        product = _find_product(products, nm_id, source["vendor_code"])
        cost_price = product.cost_price if product else None
        sold_qty = source["sold_qty"]

        before_spp = _round_money(source["before_spp"])
        after_spp = _round_money(source["after_spp"])
        sales_amount = _round_money(source["sales_amount"])
        to_pay = _round_money(source["to_pay"])

        spp_amount = _round_money(before_spp - after_spp) if before_spp is not None and after_spp is not None else None
        total_cost_price = _round_money(cost_price * sold_qty) if cost_price is not None else None
        expenses = {field: _round_money(source[field]) for field in EXPENSE_FIELDS}

        result.append(
            {
                "nm_id": nm_id,
                "vendor_code": source["vendor_code"] or (product.vendor_code if product else ""),
                "product_name": source["product_name"] or (product.name if product else ""),
                "sold_qty": sold_qty,
                "sales_amount": sales_amount,
                "before_spp": before_spp,
                "spp_amount": spp_amount,
                "after_spp": after_spp,
                "to_pay": to_pay,
                "cost_price": _round_money(cost_price),
                "total_cost_price": total_cost_price,
                **expenses,
                "tax": None,
                "profit": None,
                "profit_per_unit": None,
                "margin": None,
                "drr": None,
                "orders_qty": source.get("orders_qty", 0),
                "status": "",
                "action": "",
            }
        )

    _apply_manual_tax(result, manual_tax)
    if manual_advertising is not None:
        _apply_manual_distribution(result, manual_advertising, "advertising")
    if manual_other_expenses is not None:
        _apply_manual_distribution(result, manual_other_expenses, "other_expenses")
    for item in result:
        _recalculate_item_metrics(item)
    return result


def build_items_from_source_rows(
    db: Session,
    user_id: int,
    source_rows: list[dict],
    date_from: date | None = None,
    date_to: date | None = None,
    manual_tax: float | None = None,
    manual_advertising: float | None = None,
    manual_other_expenses: float | None = None,
) -> list[dict]:
    filtered_rows = []
    for row in source_rows or []:
        row_date = _date(row.get("date"))
        if date_from and (not row_date or row_date < date_from):
            continue
        if date_to and (not row_date or row_date > date_to):
            continue
        filtered_rows.append(row)
    return build_items(
        db, user_id, list(_group_report_rows(filtered_rows).values()),
        manual_tax=manual_tax,
        manual_advertising=manual_advertising,
        manual_other_expenses=manual_other_expenses,
    )


def validation_is_ok(validation: dict | None) -> bool:
    return bool(validation and validation.get("status") == "OK")


def recalculate_report(db: Session, user_id: int, report_id: int) -> dict | None:
    report = db.scalar(select(FinancialReport).where(FinancialReport.id == report_id, FinancialReport.user_id == user_id))
    if not report:
        return None

    items = db.scalars(
        select(FinancialReportItem).where(FinancialReportItem.report_id == report.id).order_by(FinancialReportItem.id)
    ).all()
    products = _product_lookup(db, user_id)

    for item in items:
        product = _find_product(products, item.nm_id, item.vendor_code)
        if product and product.cost_price is not None:
            item.cost_price = product.cost_price
        item.total_cost_price = _round_money(item.cost_price * item.sold_qty) if item.cost_price is not None else None

    _apply_manual_tax(items, report.manual_tax)
    if report.manual_advertising is not None:
        _apply_manual_distribution(items, report.manual_advertising, "advertising")
    if report.manual_other_expenses is not None:
        _apply_manual_distribution(items, report.manual_other_expenses, "other_expenses")
    for item in items:
        _recalculate_item_metrics(item)

    db.commit()
    return report_payload(report, [_item_payload(item) for item in items])


def update_report_item_cost(db: Session, user_id: int, report_id: int, item_id: int, cost_price: float | None) -> dict | None:
    report = db.scalar(select(FinancialReport).where(FinancialReport.id == report_id, FinancialReport.user_id == user_id))
    if not report:
        return None

    target = db.scalar(
        select(FinancialReportItem).where(
            FinancialReportItem.id == item_id,
            FinancialReportItem.report_id == report.id,
            FinancialReportItem.user_id == user_id,
        )
    )
    if not target:
        return None

    next_cost = _round_money(cost_price) if cost_price is not None else None
    target.cost_price = next_cost
    target.total_cost_price = _round_money(next_cost * target.sold_qty) if next_cost is not None else None

    product = _find_product(_product_lookup(db, user_id), target.nm_id, target.vendor_code)
    if product:
        product.cost_price = next_cost

    items = db.scalars(
        select(FinancialReportItem).where(FinancialReportItem.report_id == report.id).order_by(
            FinancialReportItem.profit.asc().nullsfirst()
        )
    ).all()
    _apply_manual_tax(items, report.manual_tax)
    if report.manual_advertising is not None:
        _apply_manual_distribution(items, report.manual_advertising, "advertising")
    if report.manual_other_expenses is not None:
        _apply_manual_distribution(items, report.manual_other_expenses, "other_expenses")
    for item in items:
        _recalculate_item_metrics(item)

    db.commit()
    return report_payload(report, [_item_payload(item) for item in items])


def update_report_tax(db: Session, user_id: int, report_id: int, manual_tax: float | None) -> dict | None:
    report = db.scalar(select(FinancialReport).where(FinancialReport.id == report_id, FinancialReport.user_id == user_id))
    if not report:
        return None
    report.manual_tax = _round_money(manual_tax) if manual_tax is not None else None
    db.flush()
    return recalculate_report(db, user_id, report_id)


def latest_report(db: Session, user_id: int) -> dict | None:
    report = db.scalar(select(FinancialReport).where(FinancialReport.user_id == user_id).order_by(FinancialReport.id.desc()))
    if not report:
        return None
    if not report.validation_json:
        return report_payload(report, [])
    items = db.scalars(
        select(FinancialReportItem).where(FinancialReportItem.report_id == report.id).order_by(
            FinancialReportItem.profit.asc().nullsfirst()
        )
    ).all()
    return report_payload(report, [_item_payload(item) for item in items])


def apply_report_settings(
    db: Session,
    user_id: int,
    report_id: int,
    global_cost_price: float | None,
    tax_percent: float | None,
    global_advertising: float | None,
    global_other_expenses: float | None,
) -> dict | None:
    report = db.scalar(select(FinancialReport).where(FinancialReport.id == report_id, FinancialReport.user_id == user_id))
    if not report:
        return None

    items = db.scalars(
        select(FinancialReportItem).where(FinancialReportItem.report_id == report.id).order_by(FinancialReportItem.id)
    ).all()

    if global_cost_price is not None:
        products = _product_lookup(db, user_id)
        new_cost = _round_money(global_cost_price)
        for item in items:
            item.cost_price = new_cost
            item.total_cost_price = _round_money(new_cost * (item.sold_qty or 0)) if new_cost is not None else None
            product = _find_product(products, item.nm_id, item.vendor_code)
            if product:
                product.cost_price = new_cost

    if tax_percent is not None:
        report.manual_tax = _round_money(tax_percent)

    if global_advertising is not None:
        report.manual_advertising = _round_money(global_advertising)

    if global_other_expenses is not None:
        report.manual_other_expenses = _round_money(global_other_expenses)

    db.flush()

    _apply_manual_tax(items, report.manual_tax)
    if report.manual_advertising is not None:
        _apply_manual_distribution(items, report.manual_advertising, "advertising")
    if report.manual_other_expenses is not None:
        _apply_manual_distribution(items, report.manual_other_expenses, "other_expenses")
    for item in items:
        _recalculate_item_metrics(item)

    db.commit()
    return report_payload(report, [_item_payload(item) for item in items])


def clear_latest_report(db: Session, user_id: int) -> bool:
    report = db.scalar(select(FinancialReport).where(FinancialReport.user_id == user_id).order_by(FinancialReport.id.desc()))
    if not report:
        return False
    db.execute(delete(FinancialReportItem).where(FinancialReportItem.report_id == report.id))
    db.delete(report)
    db.commit()
    return True


def report_payload(report: FinancialReport, items: list[dict]) -> dict:
    missing_costs = sum(1 for item in items if _value(item, "cost_price") is None)
    source_rows = report.source_rows_json or []
    columns = [
        {"key": field, "label": REPORT_COLUMN_ALIASES[field][0]}
        for field in REPORT_COLUMN_ALIASES
    ]
    validation_status = report.validation_json.get("status") if report.validation_json else None
    return {
        "report": {
            "id": report.id,
            "file_name": report.file_name,
            "period_start": report.period_start.isoformat() if report.period_start else None,
            "period_end": report.period_end.isoformat() if report.period_end else None,
            "rows_count": report.rows_count,
            "products_count": report.products_count,
            "manual_tax": report.manual_tax,
            "tax_percent": report.manual_tax,
            "manual_advertising": report.manual_advertising,
            "manual_other_expenses": report.manual_other_expenses,
            "created_at": report.created_at.isoformat() if report.created_at else None,
            "status": "processed",
            "missing_costs": missing_costs,
            "source": "file",
            "validation_status": validation_status,
        },
        "columns": columns,
        "raw_rows": source_rows,
        "items": items,
        "validation": report.validation_json,
        "message": _report_message(report, missing_costs),
    }


def _report_message(report: FinancialReport, missing_costs: int) -> str:
    if not report.validation_json:
        return "Загрузите финансовый отчёт WB заново, чтобы выполнить проверку точности новым парсером."
    if missing_costs:
        return "Некоторые товары без себестоимости. Укажите себестоимость, чтобы рассчитать точную прибыль."
    return "Финансовый отчёт обработан."


def _item_payload(item: FinancialReportItem) -> dict:
    data = {
        column.name: getattr(item, column.name)
        for column in FinancialReportItem.__table__.columns
        if column.name not in {"user_id", "report_id"}
    }
    sold = data.get("sold_qty") or 0
    orders = data.get("orders_qty") or 0
    data["buyout_rate"] = round(sold / orders * 100, 1) if orders > 0 else None
    return data


def _apply_manual_tax(items: list[Any], manual_tax: float | None) -> None:
    """Apply tax rate percentage (e.g. 22.0 = 22%) per item: tax = to_pay * rate / 100."""
    for item in items:
        if manual_tax is None:
            _assign(item, "tax", None)
        else:
            to_pay = float(_value(item, "to_pay") or 0.0)
            tax = _round_money(max(to_pay, 0.0) * float(manual_tax) / 100.0)
            _assign(item, "tax", tax)


def _apply_manual_distribution(items: list[Any], total_amount: float, field: str) -> None:
    """Distribute total_amount proportionally by to_pay across items."""
    if not items:
        return
    weights = [max(float(_value(item, "to_pay") or 0.0), 0.0) for item in items]
    total_weight = sum(weights)
    if total_weight <= 0:
        equal = _round_money(total_amount / len(items))
        for item in items:
            _assign(item, field, equal)
        return
    distributed = 0.0
    for index, (item, weight) in enumerate(zip(items, weights)):
        if index == len(items) - 1:
            piece = _round_money(total_amount - distributed)
        else:
            piece = _round_money(total_amount * weight / total_weight)
            distributed += piece or 0.0
        _assign(item, field, piece)


def _recalculate_item_metrics(item: Any) -> None:
    cost_price = _value(item, "cost_price")
    sold_qty = _value(item, "sold_qty") or 0
    to_pay = _value(item, "to_pay")
    total_cost_price = _value(item, "total_cost_price")

    if cost_price is not None:
        total_cost_price = _round_money(cost_price * sold_qty)
        _assign(item, "total_cost_price", total_cost_price)

    profit = None
    profit_per_unit = None
    margin = None
    drr = _percent(_value(item, "advertising"), _value(item, "after_spp"))

    if cost_price is not None and to_pay is not None:
        profit = _round_money(
            to_pay
            - (total_cost_price or 0)
            - (_value(item, "logistics") or 0)
            - (_value(item, "storage") or 0)
            - (_value(item, "acquiring") or 0)
            - (_value(item, "spa") or 0)
            - (_value(item, "advertising") or 0)
            - (_value(item, "tax") or 0)
            - (_value(item, "penalties") or 0)
            - (_value(item, "deductions") or 0)
            - (_value(item, "other_expenses") or 0)
        )
        profit_per_unit = _round_money(profit / sold_qty) if sold_qty else None
        margin = _percent(profit, to_pay)

    _assign(item, "profit", profit)
    _assign(item, "profit_per_unit", profit_per_unit)
    _assign(item, "margin", margin)
    _assign(item, "drr", drr)
    wb_reference_expenses = sum((_value(item, field) or 0) for field in ("commission", "logistics", "storage", "penalties", "deductions", "acquiring", "spa"))
    after_spp = _value(item, "after_spp")
    sold_qty_val = _value(item, "sold_qty") or 0
    status, action = _status_action(cost_price, profit, margin, to_pay, wb_reference_expenses, after_spp, sold_qty_val)
    _assign(item, "status", status)
    _assign(item, "action", action)


def _value(item: Any, field: str) -> Any:
    return item.get(field) if isinstance(item, dict) else getattr(item, field)


def _assign(item: Any, field: str, value: Any) -> None:
    if isinstance(item, dict):
        item[field] = value
    else:
        setattr(item, field, value)


def _number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if result > 0 else None


def _read_rows(file_name: str, content: bytes) -> list[list[Any]]:
    suffix = Path(file_name).suffix.lower()
    try:
        if suffix == ".zip":
            return _read_zip(content)
        return _dispatch_reader(suffix, content)
    except FinancialReportError:
        raise
    except Exception as exc:
        raise UnreadableFinancialReport("Не удалось прочитать файл. Проверьте формат Excel/CSV.") from exc


def _dispatch_reader(suffix: str, content: bytes) -> list[list[Any]]:
    if suffix == ".csv":
        return _read_csv(content)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(content)
    if suffix == ".xls":
        return _read_xls(content)
    raise UnreadableFinancialReport("Не удалось прочитать файл. Проверьте формат Excel/CSV.")


def _read_zip(content: bytes) -> list[list[Any]]:
    import zipfile

    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise UnreadableFinancialReport("Не удалось открыть ZIP-архив.") from exc

    priority = (".xlsx", ".xlsm", ".xls", ".csv")
    with zf:
        names = zf.namelist()
        chosen = next(
            (
                name
                for ext in priority
                for name in names
                if not name.startswith("__MACOSX") and name.lower().endswith(ext)
            ),
            None,
        )
        if not chosen:
            raise UnreadableFinancialReport("В ZIP-архиве не найден файл Excel или CSV.")
        return _dispatch_reader(Path(chosen).suffix.lower(), zf.read(chosen))


def _read_csv(content: bytes) -> list[list[Any]]:
    text = None
    for encoding in ("utf-8-sig", "cp1251", "utf-16"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UnreadableFinancialReport("Не удалось прочитать файл. Проверьте формат Excel/CSV.")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(content: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UnreadableFinancialReport("Не удалось прочитать файл. Установите openpyxl.") from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    best_rows: list[list[Any]] = []
    for sheet in workbook.worksheets:
        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        if len(rows) > len(best_rows):
            best_rows = rows
    return best_rows


def _read_xls(content: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise UnreadableFinancialReport("Не удалось прочитать файл .xls. Установите xlrd.") from exc

    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]


def _columns_found(header: list[Any], column_map: dict[str, int]) -> dict[str, str | None]:
    return {
        field: _text(_cell(header, column_map.get(field))) if field in column_map else None
        for field in REPORT_COLUMN_ALIASES
    }


def _group_report_rows(rows: list[dict]) -> dict[tuple[int | None, str, str], dict]:
    grouped: dict[tuple[int | None, str, str], dict] = {}
    for row in rows:
        nm_id = row.get("nm_id")
        vendor_code = row.get("vendor_code") or ""
        product_name = row.get("product_name") or (f"Товар {nm_id}" if nm_id else vendor_code)
        key = (nm_id, vendor_code, product_name)
        item = grouped.setdefault(
            key,
            {
                "nm_id": nm_id,
                "vendor_code": vendor_code,
                "product_name": product_name,
                "sold_qty": 0,
                "sales_amount": 0.0,
                "before_spp": 0.0,
                "after_spp": 0.0,
                "to_pay": 0.0,
                "orders_qty": 0,
                **{field: 0.0 for field in EXPENSE_FIELDS},
            },
        )
        item["sold_qty"] += row.get("sold_qty") or 0
        item["sales_amount"] += row.get("sales_amount") or 0.0
        item["before_spp"] += row.get("before_spp") or 0.0
        item["after_spp"] += row.get("after_spp") or 0.0
        item["to_pay"] += row.get("to_pay") or 0.0
        item["orders_qty"] += row.get("delivery_count") or 0
        for field in EXPENSE_FIELDS:
            item[field] += row.get(field) or 0.0
    return grouped


def _build_validation(
    columns_found: dict[str, str | None],
    rows_count: int,
    period_start: date | None,
    period_end: date | None,
    source_rows: list[dict],
    grouped_items,
) -> dict:
    grouped_items = list(grouped_items)
    totals_from_excel = _validation_totals_from_rows(source_rows)
    totals_after_grouping = _validation_totals_from_items(grouped_items)
    diff = _totals_diff(totals_from_excel, totals_after_grouping)
    ok = all(abs(value) < 0.01 for value in diff.values())
    validation = {
        "columns_found": columns_found,
        "rows_count": rows_count,
        "products_count": len(grouped_items),
        "period_start": period_start.isoformat() if period_start else None,
        "period_end": period_end.isoformat() if period_end else None,
        "totals_from_excel": totals_from_excel,
        "totals_after_grouping": totals_after_grouping,
        "diff": diff,
        "status": "OK" if ok else "ERROR",
        "message": "Расчёты совпадают с финансовым отчётом WB." if ok else "Есть расхождение с отчётом WB.",
    }
    for total_key in VALIDATION_TOTALS:
        suffix = total_key.removeprefix("total_")
        validation[f"{total_key}_excel"] = totals_from_excel[total_key]
        validation[f"{total_key}_grouped"] = totals_after_grouping[total_key]
        validation[f"diff_{suffix}"] = diff[total_key]
    return validation


def _empty_validation_totals() -> dict[str, float]:
    return {field: 0.0 for field in VALIDATION_TOTALS}


def _round_totals(totals: dict[str, float]) -> dict[str, float]:
    return {field: _round_money(value) or 0.0 for field, value in totals.items()}


def _validation_totals_from_items(items) -> dict[str, float]:
    totals = _empty_validation_totals()
    for item in items:
        for total_field, item_field in VALIDATION_TOTALS.items():
            totals[total_field] += item.get(item_field) or 0.0
    return _round_totals(totals)


def _validation_totals_from_rows(rows: list[dict]) -> dict[str, float]:
    totals = _empty_validation_totals()
    for row in rows:
        for total_field, item_field in VALIDATION_TOTALS.items():
            totals[total_field] += row.get(item_field) or 0.0
    return _round_totals(totals)


def _totals_diff(excel: dict[str, float], grouped: dict[str, float]) -> dict[str, float]:
    return {
        field: _round_money((grouped.get(field) or 0.0) - (excel.get(field) or 0.0)) or 0.0
        for field in VALIDATION_TOTALS
    }


def _detect_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best_index = -1
    best_map: dict[str, int] = {}
    best_score = 0

    for index, row in enumerate(rows[:30]):
        current: dict[str, int] = {}
        for column_index, cell in enumerate(row):
            header = _normalize_header(cell)
            if not header:
                continue
            for field, aliases in NORMALIZED_ALIASES.items():
                if field not in current and _header_matches(header, aliases):
                    current[field] = column_index
        score = len(current)
        if score > best_score:
            best_index = index
            best_map = current
            best_score = score

    if best_score < 2:
        raise MissingColumns("Не найдены нужные колонки финансового отчёта WB.")
    return best_index, best_map


def _header_matches(header: str, aliases: set[str]) -> bool:
    if header in aliases:
        return True
    return any(len(alias) > 12 and (alias in header or header in alias) for alias in aliases)


def _blank_row(row: list[Any]) -> bool:
    return not any(str(cell or "").strip() for cell in row)


def _cell(row: list[Any], index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _int(value: Any) -> int | None:
    number = _money_value(value)
    if number is None:
        return None
    return int(round(number))


def _money_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    is_negative = text.startswith("(") and text.endswith(")")
    text = text.replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return -abs(number) if is_negative else number


def _is_sale_operation(operation: str, doc_type: str = "") -> bool:
    for text in (doc_type, operation):
        normalized = str(text or "").strip().lower().replace("ё", "е")
        if normalized and "продаж" in normalized and "возврат" not in normalized:
            return True
    return False


def _date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return date.fromordinal(date(1899, 12, 30).toordinal() + int(value))
        except ValueError:
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _upsert_products_from_items(db: Session, user_id: int, parsed_items: list[dict]) -> None:
    """Register products from financial report without overwriting cost_price."""
    token = db.scalar(
        select(WbToken)
        .where(WbToken.user_id == user_id, WbToken.is_active.is_(True), WbToken.token_status != "invalid")
        .order_by(WbToken.id.desc())
    )
    existing = {
        p.nm_id: p
        for p in db.scalars(select(Product).where(Product.user_id == user_id)).all()
    }
    for item in parsed_items:
        nm_id = item.get("nm_id")
        if not nm_id:
            continue
        vendor_code = item.get("vendor_code") or ""
        name = item.get("product_name") or ""
        if nm_id in existing:
            product = existing[nm_id]
            if vendor_code and not product.vendor_code:
                product.vendor_code = vendor_code
            if name and not product.name:
                product.name = name
        else:
            product = Product(
                user_id=user_id,
                wb_token_id=token.id if token else None,
                nm_id=nm_id,
                vendor_code=vendor_code,
                name=name,
            )
            db.add(product)
            existing[nm_id] = product
    db.flush()


def _product_lookup(db: Session, user_id: int) -> dict[str, dict]:
    all_products = db.scalars(
        select(Product).where(Product.user_id == user_id).order_by(Product.id.desc())
    ).all()
    by_nm: dict[int, Product] = {}
    by_vendor: dict[str, Product] = {}
    for product in all_products:
        if product.nm_id not in by_nm:
            by_nm[product.nm_id] = product
        if product.vendor_code and product.vendor_code not in by_vendor:
            by_vendor[product.vendor_code] = product
    return {"by_nm": by_nm, "by_vendor": by_vendor}


def _find_product(products: dict[str, dict], nm_id: int | None, vendor_code: str):
    if nm_id and nm_id in products["by_nm"]:
        return products["by_nm"][nm_id]
    if vendor_code and vendor_code in products["by_vendor"]:
        return products["by_vendor"][vendor_code]
    return None


def _status_action(
    cost_price: float | None,
    profit: float | None,
    margin: float | None,
    to_pay: float | None,
    wb_reference_expenses: float = 0.0,
    after_spp: float | None = None,
    sold_qty: int = 0,
) -> tuple[str, str]:
    if sold_qty == 0:
        if wb_reference_expenses > 0:
            return "Расходы без продаж", "Проверить списания WB"
        return "Нет продаж", "Нет действия"
    if cost_price is None:
        return "Нет себестоимости", "Указать себестоимость"
    if profit is None or to_pay is None:
        return "Нет данных", "Проверить отчёт"
    if profit < 0:
        return "В минусе", "Проверить цену и расходы"
    if margin is not None and margin < 15:
        return "Низкая маржа", "Поднять цену"
    base = abs(after_spp) if after_spp else abs(to_pay)
    if base and wb_reference_expenses / base * 100 > 50:
        return "Высокие расходы WB", "Проверить логистику"
    return "В плюсе", "Контролировать остатки"


def _round_money(value: float | None) -> float | None:
    return None if value is None else round(float(value), 2)


def _percent(part: float | None, total: float | None) -> float | None:
    if part is None or not total:
        return None
    return round(part / total * 100, 1)
