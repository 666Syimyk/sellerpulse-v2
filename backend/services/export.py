"""
Генерация Excel-отчёта по данным дашборда.
Использует openpyxl (уже в requirements.txt).
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT = Font(bold=True)

COLUMNS = [
    ("Арт. продавца", "vendor_code", 18),
    ("Наименование", "name", 30),
    ("Бренд", "brand", 15),
    ("Категория", "category", 18),
    ("NM ID", "nm_id", 12),
    ("Себест. ед.", "unit_cost_price", 13),
    ("Продажи, шт.", "sold_qty", 13),
    ("Выручка до СПП", "before_spp", 16),
    ("СПП", "spp", 12),
    ("К выплате", "after_spp", 14),
    ("Себест. итого", "cost_price", 14),
    ("Комиссия", "commission", 13),
    ("Логистика", "logistics", 13),
    ("Хранение", "storage", 13),
    ("Возвраты", "returns", 13),
    ("Эквайринг", "acquiring", 13),
    ("СПП расход", "spa", 13),
    ("Реклама", "advertising", 13),
    ("Налог", "tax", 12),
    ("Штрафы", "penalties", 13),
    ("Удержания", "deductions", 13),
    ("Прочие расходы", "other_expenses", 15),
    ("Прибыль", "profit", 14),
    ("Прибыль/ед.", "profit_per_unit", 13),
    ("Маржа %", "margin", 12),
    ("ДРР %", "drr", 10),
    ("Остаток, шт.", "stock", 13),
    ("Дней запасов", "days_left", 14),
    ("Статус", "status", 16),
    ("Точность", "data_accuracy", 18),
]

METRICS_LABELS = [
    ("Период", "period"),
    ("Продажи, шт.", "sold_qty"),
    ("Выручка", "sales_sum"),
    ("К выплате", "after_spp"),
    ("Прибыль", "net_profit"),
    ("Маржа %", "margin"),
    ("ДРР %", "drr"),
    ("Возвраты, шт.", "returns_qty"),
    ("Выкуп %", "buyout_percent"),
]


def _val(v):
    if v is None:
        return ""
    return v


def generate_excel(dashboard_data: dict) -> bytes:
    wb = Workbook()

    # ── Лист 1: Сводка ────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    metrics = dashboard_data.get("metrics", {})
    shop = dashboard_data.get("shop", {})

    summary_rows = [
        ("Магазин", shop.get("name") or "—"),
        ("Период", dashboard_data.get("period", "")),
        ("", ""),
        *[(label, _val(metrics.get(key))) for label, key in METRICS_LABELS[1:]],
    ]

    ws_summary["A1"] = "SellerPulse — Сводный отчёт"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")

    for i, (label, value) in enumerate(summary_rows, start=3):
        ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 22

    # ── Лист 2: Товары ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Товары")

    headers = [col[0] for col in COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    products = dashboard_data.get("products", [])
    for row_idx, product in enumerate(products, start=2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            val = product.get(key)
            ws.cell(row=row_idx, column=col_idx, value=_val(val))

    # Итоговая строка
    total_row = len(products) + 2
    ws.cell(row=total_row, column=1, value="ИТОГО").font = _TOTAL_FONT
    summary_keys = {"sold_qty", "before_spp", "spp", "after_spp", "cost_price",
                    "commission", "logistics", "storage", "returns", "acquiring",
                    "spa", "advertising", "tax", "penalties", "deductions",
                    "other_expenses", "profit"}
    for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
        if key in summary_keys:
            values = [p.get(key) for p in products if p.get(key) is not None]
            if values:
                cell = ws.cell(row=total_row, column=col_idx, value=sum(values))
                cell.fill = _TOTAL_FILL
                cell.font = _TOTAL_FONT

    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
